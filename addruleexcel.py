import os
import argparse
import sys
import uuid
import json
from openpyxl import load_workbook
from pyakamai import pyakamai

jobId = str(uuid.uuid1())
logfilepath = ''
excelFilename = "data.xlsx"
excelTab = "MasterSheet"
configJson = {}


def getBasicAuthHeader(username, password):
    import base64
    credentials = f'{username}:{password}'
    encoded_credentials = base64.b64encode(credentials.encode('utf-8')).decode('utf-8')
    encoded_credentialsstr = 'Basic {}'.format(encoded_credentials)
    return encoded_credentialsstr


def getRule(hostname, username, password):
    f = open('template.json', 'r')
    authrule = json.load(f)
    authrule['name'] = "Auth - {}".format(hostname)
    authrule['criteria'][0]['options']['values'] = [hostname]
    authrule['children'][0]['criteria'][1]['options']['values'] = [getBasicAuthHeader(username, password)]
    return authrule


def prepareData(startRow, endRow):
    startRow = int(startRow)
    endRow = int(endRow)
    if startRow <= 0 or endRow <= 0 or startRow > endRow:
        print("Invalid Values for startRow or endRow. Please try again!", file=sys.stderr)
        exit(1)

    try:
        workbook = load_workbook(excelFilename)
        sheet_instance = workbook[excelTab]
    except Exception as e:
        print(f"Error: {e}. Please provide the correct Excel filename and tab name and try again!", file=sys.stderr)
        exit(2)

    for row in sheet_instance.iter_rows(min_row=startRow, max_row=endRow, values_only=True):
        print(row)
        item = {}
        item['HostName'] = row[1]
        item['Username'] = row[2]
        item['Password'] = row[3]

        if row[0] not in configJson:
            configJson[row[0]] = [item]
        else:
            configJson[row[0]].append(item)
        print(item)

    print("Data Preparation Done!")
    print(json.dumps(configJson, indent=2))


def addRule(ChangeID, accountSwitchKey):
    for key in configJson.keys():
        config = key
        pyakamaiObj = pyakamai(accountSwitchKey=accountSwitchKey,edgercLocation='~/.edgerc', section='default',debug=False, verbose=False)
        akamaiconfig = pyakamaiObj.client('property')
        akamaiconfig.config(config)
        newversion = akamaiconfig.createVersion(akamaiconfig.getProductionVersion())
        ruleTree = akamaiconfig.getRuleTree(newversion)
        for hosts in configJson[key]:
            childrule = getRule(hosts['HostName'], hosts['Username'], hosts['Password'])
            print("Preparing the rule for {} for the config {}".format(hosts['HostName'], config))
            ruleTree['rules']['children'].append(childrule)

        ruleTree['ruleFormat'] = 'v2023-10-30'
        propruleInfo_json = json.dumps(ruleTree, indent=2)
        print("Updating the rules for the config {}".format(config))
        addRuleStatus = akamaiconfig.updateRuleTree(newversion, propruleInfo_json)
        updateNoteStatus = akamaiconfig.addVersionNotes(newversion, ChangeID)
        if addRuleStatus and updateNoteStatus:
            print("Updated the Rule successfully for the config {}".format(config))
            activationStatus = akamaiconfig.activateStaging(newversion, ChangeID, ["apadmana@akamai.com"])
            if activationStatus:
                print("Initiated the activation of the config {}".format(config))
            else:
                print("Failed to activate the config {}".format(config))
        else:
            print("Failed to update the Rule successfully for the config {}".format(config))
        print('*' * 80)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Times CDN Onboarding Tool.')
    # Storage migration
    parser.add_argument('--start', required=True, help='Starting Row Number')
    parser.add_argument('--end', required=True, help='End Row')
    parser.add_argument('--accountSwitchKey', default=None, help='Account SwitchKey')
    parser.add_argument('--ChangeID', required=True, help='ChangeID')

    args = parser.parse_args()

    curdir = os.getcwd()
    dirpath = os.path.dirname(curdir + '/logs')
    logfilepath = dirpath + "/" + jobId + '.txt'

    # sys.stdout = open(logfilepath, 'w+')

    prepareData(args.start, args.end)
    addRule(args.ChangeID, args.accountSwitchKey)

'''
python addruleexcel.py --accountSwitchKey B-C-1IE2OH8 --ChangeID 'Adding Phase2 Basic Auth' --start 2 --end 2 
'''
